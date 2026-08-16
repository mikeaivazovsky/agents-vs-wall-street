"""Write figures into the organisers' workbook templates.

This layer knows where a figure goes and what shape it must have. It does not
know where the figure came from, how it was calculated or how confident anyone
is in it. It receives a figure, checks it against the specification for the
target cell and writes it.

The templates carry the structure the organisers check: the sheet name, the
metric labels, the units column and the period header. None of that is created
here. A template is opened, the input cells are filled and the result is saved
under a new name, so the structure arrives unchanged.
"""

from __future__ import annotations

from pathlib import Path
from types import TracebackType

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook

from src.domain.domain_values import Figure, UnitMismatch
from src.domain.spec import MetricSpec

# The templates place the metric label and the unit to the left of the input
# cell on the same row. Reading them back is how the writer confirms that it is
# filling the cell the specification meant.
_LABEL_COLUMN = 1
_UNIT_COLUMN = 2


class WorkbookError(RuntimeError):
    """Raised when a workbook cannot be opened, filled or saved."""


class TemplateMismatch(WorkbookError):
    """Raised when a template does not match the specification.

    This means the specification and the supplied template have drifted apart.
    Writing anyway would put a figure in a cell whose meaning is unknown.
    """


class WorkbookWriter:
    """Fills one workbook and saves it.

    A workbook holds three figures, so the template is opened once, written to
    three times and saved once. Use it as a context manager, or call close to
    release the file if the run is abandoned partway.
    """

    def __init__(self, template_path: Path, output_path: Path) -> None:
        self._template_path = template_path
        self._output_path = output_path
        self._workbook: Workbook | None = None
        self._written: list[str] = []

    def __enter__(self) -> "WorkbookWriter":
        self.open()
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        self.close()

    @property
    def output_path(self) -> Path:
        return self._output_path

    @property
    def written_cells(self) -> tuple[str, ...]:
        return tuple(self._written)

    def open(self) -> None:
        """Load the template into memory."""
        if self._workbook is not None:
            return
        try:
            self._workbook = load_workbook(self._template_path)
        except FileNotFoundError as error:
            raise WorkbookError(
                f"template {self._template_path.name} is missing"
            ) from error
        except (InvalidFileException, OSError) as error:
            raise WorkbookError(
                f"template {self._template_path.name} cannot be opened: {error}"
            ) from error

    def close(self) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def write(self, metric: MetricSpec, figure: Figure) -> None:
        """Put one figure into the cell the specification names for it.

        The figure must already be in the unit the specification declares.
        Nothing is converted here: a figure in the wrong unit is a fault
        upstream, and quietly rescaling it would turn a visible failure into a
        wrong number in the submitted workbook.
        """
        workbook = self._require_open()
        sheet = self._require_sheet(workbook, metric)
        self._verify_row(sheet, metric)

        try:
            figure.require_unit(metric.unit)
        except UnitMismatch as error:
            raise TemplateMismatch(
                f"{metric.id.value} into {metric.workbook}!{metric.cell}: {error}"
            ) from error

        # Assigning to the cell replaces its value and leaves the template's
        # own formatting in place, so the input cell keeps the fill and the
        # number format the organisers applied to it.
        sheet[metric.cell] = float(figure.value)
        self._written.append(metric.cell)

    def save(self) -> Path:
        """Write the filled workbook to its output path."""
        workbook = self._require_open()
        try:
            self._output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self._output_path)
        except OSError as error:
            raise WorkbookError(
                f"cannot save {self._output_path.name}: {error.strerror}"
            ) from error
        return self._output_path

    def _require_open(self) -> Workbook:
        if self._workbook is None:
            raise WorkbookError("the workbook is not open")
        return self._workbook

    @staticmethod
    def _require_sheet(workbook: Workbook, metric: MetricSpec):
        if metric.sheet not in workbook.sheetnames:
            raise TemplateMismatch(
                f"{metric.workbook} has no sheet named {metric.sheet!r}"
            )
        return workbook[metric.sheet]

    @staticmethod
    def _verify_row(sheet, metric: MetricSpec) -> None:
        """Check that the target row still holds the metric it should.

        The specification records a cell address, and an address only means
        something while the template's layout stays put. Reading back the label
        and the unit beside the target cell turns a silent misplacement into a
        clear failure before anything is written.
        """
        row, _ = coordinate_to_tuple(metric.cell)

        found_label = sheet.cell(row=row, column=_LABEL_COLUMN).value
        if _text(found_label) != metric.label:
            raise TemplateMismatch(
                f"{metric.workbook}!{metric.sheet} row {row} should be labelled "
                f"{metric.label!r} but holds {_text(found_label)!r}"
            )

        found_unit = sheet.cell(row=row, column=_UNIT_COLUMN).value
        if _text(found_unit) != metric.unit.value:
            raise TemplateMismatch(
                f"{metric.workbook}!{metric.sheet} row {row} should be in "
                f"{metric.unit.value!r} but the template says {_text(found_unit)!r}"
            )


def _text(value: object) -> str:
    """Cell text with surrounding space removed, empty for a blank cell."""
    return "" if value is None else str(value).strip()
